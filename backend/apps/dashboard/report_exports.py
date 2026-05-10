from io import BytesIO
from pathlib import Path

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


FONT_NAME = 'ReportSans'


def _register_cyrillic_font():
    if FONT_NAME in pdfmetrics.getRegisteredFontNames():
        return FONT_NAME

    candidates = [
        Path('C:/Windows/Fonts/arial.ttf'),
        Path('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'),
        Path('/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf'),
    ]
    for path in candidates:
        if path.exists():
            pdfmetrics.registerFont(TTFont(FONT_NAME, str(path)))
            return FONT_NAME
    return 'Helvetica'


def build_csv_response(filename, headers, rows):
    import csv

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def build_xlsx_response(filename, sheet_title, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title[:31]
    sheet.append(headers)

    header_fill = PatternFill(fill_type='solid', fgColor='1F2328')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        sheet.append([value if isinstance(value, (int, float)) or value is None else str(value) for value in row])

    for column in sheet.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, 12), 42)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def build_pdf_response(filename, title, headers, rows):
    font_name = _register_cyrillic_font()
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    styles['Title'].fontName = font_name
    styles['Normal'].fontName = font_name

    table_data = [headers] + [[str(value or '') for value in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2328')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cfd4dc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f6f7f9')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('LEADING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))

    document.build([Paragraph(title, styles['Title']), Spacer(1, 6 * mm), table])
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response
